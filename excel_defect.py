#!/usr/bin/env python3
# coding: utf-8

import numpy as np
import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from pathlib import Path
from openpyxl.utils.dataframe import dataframe_to_rows
import glob
import pickle
glob.glob('/pycharm/*') 


def comparison(word):
	regexp =re.finditer(r'\w+=[0-9]+',word)
	regexp_minus_f = re.finditer(r'-[0-9]+.[0-9]+',word)
	regexp_minus = re.finditer(r'-[0-9]+(^.)',word)

	for m in regexp:
		matched_word=m.group()
		new_word="==".join(matched_word.split("="))
		word=word.replace(matched_word, new_word)
	return word


def import_excel(fileName,sheetName,title,dataframe):
	book = openpyxl.Workbook()
	book.create_sheet('tt_results')#don't give extension
	# Acquire a sheet by its name
	tt_results = book.get_sheet_by_name('tt_results')
	# Writing to sheet
	rows = dataframe_to_rows(tt_df)
	for r_idx, row in enumerate(rows, 3):
		for c_idx, value in enumerate(row, 1):
			 ws.cell(row=r_idx, column=c_idx, value=value)
	tt_results.cell(row=1, column=1).value = title
	book.save('Sample.xlsx')


def print_df(df_name,row_pos,col_pos,sheet_name):
	rows = dataframe_to_rows(df_name)
	
	for r_idx, row in enumerate(rows, row_pos):
		for c_idx, value in enumerate(row, col_pos):
			 sheet_name.cell(row=r_idx, column=c_idx, value=value)
	

def checkSheet(file_handler,sheet_name):
	if sheet_name in file_handler.sheetnames:
		sheet_name = file_handler.get_sheet_by_name(str(sheet_name))
	else:
		file_handler.create_sheet(sheet_name)#don't give extension
		sheet_name = file_handler.get_sheet_by_name(str(sheet_name))
	return sheet_name


def checkFile(path,filename):
	my_file = Path(path+filename)
	if my_file.is_file():
		book = load_workbook(filename)
	else:
		book = Workbook()
	return book


def printExcel(path,fileName,case_df,tt_df,tf_df,ft_df,ff_df):
	book=checkFile(path,fileName+'.xlsx')

	tt_results =checkSheet(book,'tt_results')
	tf_results =checkSheet(book,'tf_results')
	ft_results =checkSheet(book,'ft_results')
	ff_results =checkSheet(book,'ff_results')
	print_df(case_df,1,1,ff_results)
	print_df(ff_df,5,1,ff_results)

	book.save(path+'/'+fileName+'.xlsx')


def models_pickle(path_data):
	model_load = []
	os.chdir(path_data + "/models")
	for file in glob.glob("*.pkl"):
		model_load.append(file)
	return model_load


def generate_excels():
	datasets_list = ['hive','groovy','activemq','derby','camel','lucene','hbase','wicket','jruby']

	for dataset_kw in datasets_list:
		print(dataset_kw,"dataset_kw")
		path_data = "/home/dilini/defects/ICSEnew/"+dataset_kw
		path = "/home/dilini/defects/ICSEnew/"+dataset_kw+"/data_MO/"
		all_files = glob.glob(os.path.join(path, dataset_kw+"_*.csv")) #make list of paths
		data_test = pd.read_csv("/home/dilini/defects/ICSEnew/"+dataset_kw+"/datasets/"+dataset_kw+"-R2.csv", skipinitialspace=True)

		for file in all_files:
			model_load = models_pickle(path_data)
			
			for nb, black in enumerate(model_load):
				with open(path_data + '/models/' + black, 'rb') as f:
					blackbox = pickle.load(f)
				model_name = str(black).split(".")
				model_name = str(model_name[0])

				# Getting the file name without extension
				file_name = os.path.splitext(os.path.basename(file))[0]
				case_name = int(file_name.split("_")[1])
				model_name_c = str(file_name.split("_")[4])

				if model_name == model_name_c:

					tt_df= pd.DataFrame([])
					tf_df= pd.DataFrame([])
					ft_df= pd.DataFrame([])
					ff_df= pd.DataFrame([])

					case_data = data_test.iloc[case_name].to_frame().T 
					case_data.drop('Unnamed: 0', axis=1, inplace=True)
					real_target = blackbox.predict(case_data.iloc[:, case_data.columns != 'target'])
					real_target= 'target=='+str(real_target[0])+str('.000')

					# Reading the file content to create a DataFrame
					rules_df = pd.read_csv(file)#rule file
					for  index, row in rules_df.iterrows():
						rule = rules_df.iloc[index,0]
						rule = comparison(rule)
						
						class_val = rules_df.iloc[index,1]
						class_val = comparison(class_val)

						if real_target==class_val:
							print("correctly predicted")
						#Risky practices that lead a model to predict a fileas  defective
						if case_data.eval(rule).all()==True and real_target==class_val :
							tt_df=tt_df.append(row.to_frame().T)
						# Non-risky practices that lead a model to predict afile  as  clean
						elif case_data.eval(rule).all()==True and real_target!=class_val:
							tf_df=tf_df.append(row.to_frame().T)
						#Practices to avoid to not increase the risk of havingdefects
						elif case_data.eval(rule).all()==False and real_target==class_val:
							ft_df=ft_df.append(row.to_frame().T)
						#Practices  to  follow  to  decrease  the  risk  of  having defects
						elif case_data.eval(rule).all()==False and real_target!=class_val:
							ff_df=ff_df.append(row.to_frame().T)

					printExcel(path,file_name,case_data,tt_df,tf_df,ft_df,ff_df)
					
				else:
					continue
				
	# Setting the file name (without extension) as the index name
def main():
	generate_excels()


if __name__ == '__main__':
	main()
#


